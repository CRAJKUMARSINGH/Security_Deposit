import os
from openpyxl import load_workbook

TARGET_DIR = "/workspace/Security_Refund_Sheets_2025_20250903_033335"

def fix_workbook(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        # 1) Apply thin borders for A20:B26
        from openpyxl.styles import Border, Side
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row_idx in range(20, 27):
            for col in ("A", "B"):
                ws[f"{col}{row_idx}"].border = thin

        # 2) Double the height of row 32 (approx 40 points if default ~20)
        ws.row_dimensions[32].height = 40

        # 3) Remove borders in certificate section (heuristic: lines after a row that equals "Certified That:-" until signatures)
        # Find certificate start
        cert_start = None
        for r in range(1, ws.max_row + 1):
            val = ws[f"A{r}"].value
            if isinstance(val, str) and val.strip().startswith("Certified That:-"):
                cert_start = r
                break
        if cert_start:
            # Clear borders for a reasonable range (up to next 10 lines) until blank then stop
            for r in range(cert_start, min(cert_start + 12, ws.max_row + 1)):
                for c in ("A", "B", "C", "D", "E"):
                    ws[f"{c}{r}"].border = Border()

        # 4) Ensure no border at A4
        if ws["A4"].border is not None:
            ws["A4"].border = Border()

        # 5) Print: A4 portrait one page
        # Set print area to cover A1:E(last row with content)
        last_row = 0
        for r in range(1, ws.max_row + 1):
            if ws[f"A{r}"].value is not None:
                last_row = r
        if last_row:
            ws.print_area = f"A1:E{last_row+2}"
        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

    wb.save(path)


def main():
    for name in os.listdir(TARGET_DIR):
        if name.endswith('.xlsx') and not name.startswith('~$'):
            fix_workbook(os.path.join(TARGET_DIR, name))

if __name__ == '__main__':
    main()

