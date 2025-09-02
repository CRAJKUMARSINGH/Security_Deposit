import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.worksheet.hyperlink import Hyperlink
import os
from datetime import datetime

def read_excel_data(file_path, sheet_name='agency'):
    """Read data from Excel file agency sheet"""
    try:
        # First, let's see what sheets are available
        xl_file = pd.ExcelFile(file_path)
        print(f"Available sheets: {xl_file.sheet_names}")
        
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"Successfully read {len(df)} rows from {sheet_name} sheet")
        print(f"Columns: {list(df.columns)}")
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def create_sheet_name(vendor, agreement_no):
    """Create sheet name from vendor and agreement number"""
    try:
        # Get vendor name and clean it
        vendor_str = str(vendor).strip()
        
        # Remove 'M/s' prefix variations
        vendor_clean = vendor_str.replace('M/s', '').replace('M/s.', '').replace('M/s ', '').strip()
        
        # Get first word (first name)
        first_name = vendor_clean.split()[0] if vendor_clean else 'Unknown'
        
        # Get agreement number and clean it
        agreement_str = str(agreement_no).strip()
        
        # Extract number part before year (e.g., "104/2020-21" -> "104")
        if '/' in agreement_str:
            agreement_clean = agreement_str.split('/')[0]
        elif '-' in agreement_str:
            agreement_clean = agreement_str.split('-')[0]
        else:
            agreement_clean = agreement_str
        
        # Create sheet name: First name + space + agreement number
        sheet_name = f"{first_name} {agreement_clean}"
        
        # Excel sheet names cannot contain certain characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '')
        
        # Excel sheet names have 31 character limit
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        # Ensure it's not empty
        if not sheet_name.strip():
            sheet_name = f"Work_{agreement_clean}"
        
        # Final cleanup
        sheet_name = sheet_name.strip()
        
        print(f"Creating sheet: '{sheet_name}' from vendor: '{vendor_str}' and agreement: '{agreement_str}'")
        
        return sheet_name
        
    except Exception as e:
        print(f"Error creating sheet name: {e}")
        # Fallback naming
        try:
            agreement_clean = str(agreement_no).split('/')[0] if '/' in str(agreement_no) else str(agreement_no)
            return f"Work_{agreement_clean}"
        except:
            return "Work_Unknown"

def create_single_work_sheet(wb, row, work_idx):
    """Create a single work sheet with enhanced formatting"""
    
    # Create new worksheet with correct column names
    vendor_name = row['Name of Contractor'] if 'Name of Contractor' in row else ''
    agreement_no = row['Agreement No.'] if 'Agreement No.' in row else ''
    
    sheet_name = create_sheet_name(vendor_name, agreement_no)
    ws = wb.create_sheet(title=sheet_name)
    
    # Define enhanced styles
    title_font = Font(bold=True, size=16, color='000080')  # Navy blue
    header_font = Font(bold=True, size=12, color='000000')
    normal_font = Font(size=11, color='000000')
    small_font = Font(size=10, color='000000')
    value_font = Font(size=11, bold=True, color='000000')
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Enhanced borders
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header background
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    
    current_row = 1
    
    # Main title with enhanced styling - merged A to E
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = f"ORDER FOR REFUND OF SECURITY DEPOSIT [RWMF 119]"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = thick_border
    ws[f'A{current_row}'].fill = header_fill
    current_row += 1  # Only one blank row after heading
    
    # Form fields with enhanced formatting - using correct column names
    form_fields = [
        ("1. Name of Contractor:", vendor_name),
        ("2. Amount of Deposit: ₹", ""),  # Not available in current data
        ("3. Name of Work:", row['Name of Work'] if 'Name of Work' in row else ''),
        ("4. Agreement No.:", agreement_no),
        ("5. Reference for granting refunds:", ""),
        ("6. Date of Commencement:", row['Date of Commencement'] if 'Date of Commencement' in row else ''),
        ("7. Stipulated date of Completion:", row['Stipulated date of Completion'] if 'Stipulated date of Completion' in row else ''),
        ("8. Actual Date of Completion:", row['Actual Date of Completion'] if 'Actual Date of Completion' in row else ''),
        ("9. MB No.:", ""),
        ("10. Date of Payment of final bill:", ""),
        ("11. Date of Expiry of 3/6 months/DLP:", ""),
        ("12. Was work satisfactory:", "Yes"),
        ("13. Any tools outstanding against contractor:", "Nil"),
        ("14. Any recovery due from contractor after payment of final bill:", "Nil"),
        ("15. Extension of time limit sanctioned vide", ""),
        ("16. Assistant Engineer Signature's Recommending refund", ""),
        ("17. Accountant's Remarks", "")
    ]
    
    for field_label, field_value in form_fields:
        if field_label == "3. Name of Work:":
            # Name of work spans A to E
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = f"{field_label} {field_value}"
            ws[f'A{current_row}'].font = value_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        else:
            # All other items: Label in A, Value in E (shifting from C to E)
            ws[f'A{current_row}'] = field_label
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].alignment = left_alignment
            # No border for form fields A2 to A18
            
            ws[f'E{current_row}'] = field_value
            ws[f'E{current_row}'].font = value_font
            ws[f'E{current_row}'].alignment = left_alignment
            ws[f'E{current_row}'].border = thin_border
        
        current_row += 1
    
    # No blank row between item 17 and 18
    
    # Security Deposit Details Table - corrected header
    ws[f'A{current_row}'] = "18. Details of Security Deposit"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = left_alignment
    ws[f'A{current_row}'].border = thin_border
    current_row += 1
    
    # Table headers - merge A and B, shift others to C, D, E
    headers = ["Bill Type", "MB No.", "SD Type", "Amount (₹)"]
    
    # First header spans A and B
    table_header_row = current_row
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = headers[0]
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].fill = header_fill
    
    # Remaining headers in C, D, E
    for col_idx, header in enumerate(headers[1:], 1):
        col_letter = chr(67 + col_idx - 1)  # C, D, E
        ws[f'{col_letter}{current_row}'] = header
        ws[f'{col_letter}{current_row}'].font = header_font
        ws[f'{col_letter}{current_row}'].alignment = center_alignment
        ws[f'{col_letter}{current_row}'].border = thin_border
        ws[f'{col_letter}{current_row}'].fill = header_fill
    
    current_row += 1
    
    # Enhanced table data rows - A-B merged, then C, D, E
    table_data = [
        ("", "", "", ""),  # Row 1 - empty for filling
        ("", "", "", ""),  # Row 2 - empty for filling
        ("", "", "", ""),  # Row 3 - empty for filling
        ("", "", "", ""),  # Row 4 - empty for filling
        ("", "", "", ""),  # Row 5 - empty for filling
        ("Total:", "", "", "₹[Amount to be filled]")
    ]
    
    for table_row in table_data:
        # First column spans A and B
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = table_row[0]
        ws[f'A{current_row}'].font = normal_font
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        # Remaining columns in C, D, E
        for col_idx, value in enumerate(table_row[1:], 1):
            col_letter = chr(67 + col_idx - 1)  # C, D, E
            ws[f'{col_letter}{current_row}'] = value
            ws[f'{col_letter}{current_row}'].font = normal_font
            ws[f'{col_letter}{current_row}'].alignment = center_alignment
            ws[f'{col_letter}{current_row}'].border = thin_border
        current_row += 1
    
    # Ensure borders are correctly applied to A20:B26 (table header + 6 data rows)
    # Based on the current layout, this corresponds to table_header_row through table_header_row + len(table_data)
    for row_idx in range(table_header_row, table_header_row + len(table_data) + 1):
        for col_letter in ('A', 'B'):
            ws[f'{col_letter}{row_idx}'].border = thin_border

    # Explicitly enforce borders on A20:B26 regardless of layout shifts
    for row_idx in range(20, 27):
        for col_letter in ('A', 'B'):
            ws[f'{col_letter}{row_idx}'].border = thin_border
    
    # No blank rows after total
    
    # Certification section - no borders, proper formatting
    certification_items = [
        "Certified That:-",
        "1. The Work has been completed as per G-schedule.",
        "2. The work has been inspected by the undersigned as on and it stood satisfactory.",
        "3. No Defect found during DLP Period.",
        "4. The final time extension granted upto With/without compensation by the competent authority.",
        "5. The defects pointed out by higher authorities or other authorized authorities during inspection etc have been removed by the contractor and compliance has been refund."
    ]
    
    for cert_item in certification_items:
        if cert_item.startswith("Certified That:-"):
            ws[f'A{current_row}'] = cert_item
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].alignment = left_alignment
        elif cert_item.startswith("5."):
            # Special handling for point 5 - span columns A to E with wrap text and proper height
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = cert_item
            ws[f'A{current_row}'].font = small_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[current_row].height = 26  # Set to 26 points as requested
        else:
            ws[f'A{current_row}'] = cert_item
            ws[f'A{current_row}'].font = small_font
            ws[f'A{current_row}'].alignment = left_alignment
        current_row += 1
    
    # Signature section - no borders, proper layout
    current_row += 1
    
    # Signature items with PWD Electric Div.- Udaipur below Executive Engineer - no blank rows
    signature_items = [
        ("Divisional Accountant", "Assistant Engineer", "Executive Engineer"),
        ("", "", "PWD Electric Div.- Udaipur")  # Directly below Executive Engineer
    ]
    
    for sig_row in signature_items:
        for col_idx, sig_text in enumerate(sig_row):
            col_letter = chr(65 + col_idx * 2)  # A, C, E (spacing between signatures)
            if sig_text:  # Only write if there's text
                ws[f'{col_letter}{current_row}'] = sig_text
                ws[f'{col_letter}{current_row}'].font = normal_font
                ws[f'{col_letter}{current_row}'].alignment = center_alignment
                # No border for signature section
        current_row += 1
    
    # Auto-adjust column widths with better formatting for wrapped text
    column_widths = {'A': 30, 'B': 5, 'C': 25, 'D': 25, 'E': 25, 'F': 15, 'G': 15, 'H': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row heights for better appearance
    for row_num in range(1, current_row + 5):
        ws.row_dimensions[row_num].height = 20
    
    # Ensure certificate point 5 maintains its 26 point height (override general setting)
    # Find the certificate point 5 row and set its height to 26
    for row_num in range(1, current_row + 5):
        cell_value = ws[f'A{row_num}'].value
        if cell_value and str(cell_value).startswith("5."):
            ws.row_dimensions[row_num].height = 26
            break
    
    # Special case: Double the height of row 32 (~40 points)
    ws.row_dimensions[32].height = 40
    
    # Setup default print layout for all sheets
    setup_default_print_layout(ws)
    
    return ws

def setup_default_print_layout(ws):
    """Setup default print layout for all sheets to fit on 1 page"""
    
    # Set print area to cover all content (A1 to E with last row)
    # Find the last row with content
    last_row = 0
    for row_num in range(1, 50):  # Check up to row 50
        if ws[f'A{row_num}'].value is not None:
            last_row = row_num
    
    # Set print area to A1:E{last_row + 2} to ensure all content is included
    print_area = f'A1:E{last_row + 2}'
    ws.print_area = print_area
    
    # Set page setup for A4 portrait
    ws.page_setup.paperSize = 9  # A4 paper size
    ws.page_setup.orientation = 'portrait'
    
    # Set smaller margins to fit more content
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    
    # Set scaling to fit to 1 page wide and tall
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1  # Fit to 1 page tall
    
    # Center the page horizontally
    ws.print_options.horizontalCentered = True
    
    # Set header and footer
    ws.oddHeader.center.text = "Security Deposit Refund Form"
    ws.oddFooter.center.text = "Page &P of &N"

def create_security_refund_sheet(data_batch, batch_number, agreement_year=None):
    """Create a security refund workbook with 25 separate sheets, one per work"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Process each work in the batch and create a separate sheet
    for work_idx, (_, row) in enumerate(data_batch.iterrows(), 1):
        create_single_work_sheet(wb, row, work_idx)
    
    # Add VBA macro for print functionality
    add_print_macro(wb)
    
    return wb

def add_print_macro(wb):
    """Add VBA macro for print functionality"""
    
    # VBA code for print macro
    vba_code = '''
Sub PrintCurrentSheet()
    ' Print the current sheet in A4 portrait format
    Dim ws As Worksheet
    Set ws = ActiveSheet
    
    ' Set print area to all used cells
    Dim lastRow As Long
    Dim lastCol As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
    lastCol = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    
    ' Set print area
    ws.PageSetup.PrintArea = ws.Range(ws.Cells(1, 1), ws.Cells(lastRow, lastCol)).Address
    
    ' Set page setup for A4 portrait
    With ws.PageSetup
        .PaperSize = xlPaperA4
        .Orientation = xlPortrait
        .FitToPagesWide = 1
        .FitToPagesTall = 0
        .CenterHorizontally = True
        .LeftMargin = Application.InchesToPoints(0.7)
        .RightMargin = Application.InchesToPoints(0.7)
        .TopMargin = Application.InchesToPoints(0.75)
        .BottomMargin = Application.InchesToPoints(0.75)
        .HeaderMargin = Application.InchesToPoints(0.3)
        .FooterMargin = Application.InchesToPoints(0.3)
    End With
    
    ' Print the sheet
    ws.PrintOut
    
End Sub
'''
    
    # Note: openpyxl doesn't directly support VBA macros
    # The print button will work with Excel's built-in print functionality
    # Users can manually add this VBA code if they want advanced print control

def split_data_into_batches(df, batch_size=25):
    """Split dataframe into batches of specified size"""
    batches = []
    total_rows = len(df)
    
    for i in range(0, total_rows, batch_size):
        batch = df.iloc[i:i+batch_size].copy()
        batch_number = (i // batch_size) + 1
        batches.append((batch, batch_number))
    
    return batches

def get_agreement_year_from_data(df):
    """Extract agreement year from the data for naming convention"""
    try:
        # Try to extract year from Agreement No or Start Date
        if 'Agreement No' in df.columns:
            # Look for year pattern in agreement numbers
            sample_agreement = str(df['Agreement No'].iloc[0]) if not df.empty else ''
            if len(sample_agreement) >= 4:
                # Try to find 4-digit year
                for i in range(len(sample_agreement) - 3):
                    year_candidate = sample_agreement[i:i+4]
                    if year_candidate.isdigit() and 2000 <= int(year_candidate) <= 2030:
                        return year_candidate
        
        # Fallback to current year
        return datetime.now().strftime('%Y')
    except:
        return datetime.now().strftime('%Y')

def read_work_data_from_txt(file_path):
    """Read work order data from 355.txt file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # The file is all on one line, so we need to split it properly
        # Looking at the structure, each work order starts with a number followed by the work name
        
        import re
        
        # Split the content by work order numbers (1, 2, 3, etc.)
        # Each work order starts with a number followed by the work description
        work_entries = re.split(r'(?=\d+[A-Z])', content)
        
        works = []
        for entry in work_entries:
            if entry.strip():
                # Extract work number (first digits)
                work_num_match = re.match(r'^(\d+)', entry)
                if work_num_match:
                    work_num = work_num_match.group(1)
                    
                    # Extract work name (text after work number until vendor)
                    work_name_match = re.search(r'^\d+([A-Z][^A-Z]*(?=[A-Z][a-zA-Z\s&]+(?:Enterprises|Electricals|Company|Ltd|Pvt|Traders|Suppliers|Engineering|Industries|Centre|Service)))', entry)
                    work_name = work_name_match.group(1).strip() if work_name_match else entry[:50] + '...'
                    
                    # Extract vendor
                    vendor_match = re.search(r'([A-Z][a-zA-Z\s&]+(?:Enterprises|Electricals|Company|Ltd|Pvt|Traders|Suppliers|Engineering|Industries|Centre|Service))', entry)
                    vendor = vendor_match.group(1).strip() if vendor_match else ''
                    
                    # Extract WO No (5-6 digit number)
                    wo_match = re.search(r'(\d{5,6})', entry)
                    wo_no = wo_match.group(1) if wo_match else ''
                    
                    # Extract Agreement No (pattern like "71014/2022-23")
                    agreement_match = re.search(r'(\d+[\/\s]*(?:of\s+)?\d{4}-\d{2,4})', entry)
                    agreement_no = agreement_match.group(1) if agreement_match else ''
                    
                    # Extract dates (pattern like "(26/07/2022)")
                    date_matches = re.findall(r'\((\d{2}\/\d{2}\/\d{4})\)', entry)
                    start_date = date_matches[0] if len(date_matches) > 0 else ''
                    comp_date = date_matches[1] if len(date_matches) > 1 else ''
                    
                    # Extract amounts (decimal numbers)
                    amount_matches = re.findall(r'(\d+\.\d{2})', entry)
                    amount_a = amount_matches[0] if len(amount_matches) > 0 else ''
                    amount_b = amount_matches[1] if len(amount_matches) > 1 else ''
                    
                    # Extract ACD date (last date in the entry)
                    acd_match = re.search(r'(\d{2}\/\d{2}\/\d{4})(?:\s*$)', entry)
                    acd_date = acd_match.group(1) if acd_match else ''
                    
                    work_data = {
                        'S.No': work_num,
                        'WorkOrder Name': work_name,
                        'Vendor': vendor,
                        'WO No': wo_no,
                        'Agreement No': agreement_no,
                        'Start Date Serial': '',
                        'Start Date': start_date,
                        'Comp Date': comp_date,
                        'Some Date Serial': '',
                        'Amount_a': amount_a,
                        'Amount_b': amount_b,
                        'Amount_c': '',
                        'Amount_d': '',
                        'Actual date of completion ACD': acd_date
                    }
                    works.append(work_data)
        
        df = pd.DataFrame(works)
        print(f"Successfully read {len(df)} works from {file_path}")
        return df
        
    except Exception as e:
        print(f"Error reading text file: {e}")
        return None

def main():
    """Main function to process Excel file and generate security refund sheets"""
    
    excel_file = 'work_order_master.xlsx'
    
    print("Reading Excel file Work Orders...")
    df = read_excel_data(excel_file, 'Work Orders')
    
    if df is None:
        print("Failed to read Excel file. Please check the file path and sheet name.")
        return
    
    print(f"Total works found: {len(df)}")
    
    # Get agreement year for naming
    agreement_year = get_agreement_year_from_data(df)
    print(f"Using agreement year: {agreement_year}")
    
    # Split data into batches
    batches = split_data_into_batches(df, 25)
    print(f"Created {len(batches)} batches")
    
    # Create output directory with timestamp to avoid permission issues
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"Security_Refund_Sheets_{agreement_year}_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate security refund sheets for each batch
    for batch_idx, (batch_data, batch_number) in enumerate(batches, 1):
        print(f"Processing batch {batch_idx} with {len(batch_data)} works...")
        
        # Create security refund sheet
        wb = create_security_refund_sheet(batch_data, batch_idx, agreement_year)
        
        # Save the file
        filename = f"Security_Refund_Batch_{batch_idx:02d}_{agreement_year}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        
        print(f"Saved: {filepath}")
    
    print(f"\nCompleted! Generated {len(batches)} security refund workbooks in '{output_dir}' directory.")
    print("Each workbook contains:")
    print("- 25 separate sheets (one per work order)")
    print("- Sheet names: First name of contractor + agreement number")
    print("- Enhanced formatting with elegant borders and professional styling")
    print("- Default 'Satisfactory' status for security refund")
    print("- All relevant work order data")
    print("- Print-ready format with proper spacing and alignment")

if __name__ == "__main__":
    main()
